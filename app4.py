import os
import re
import json
import pandas as pd
import streamlit as st
from openai import AzureOpenAI

# Load secrets from Streamlit Cloud or .env for local development
def get_secret(key):
    """Get secret from Streamlit secrets or environment variable"""
    try:
        # Try Streamlit secrets first (for cloud deployment)
        if hasattr(st, 'secrets'):
            try:
                return st.secrets[key]
            except KeyError:
                pass
    except Exception:
        pass
    
    # Fall back to environment variable (for local development)
    return os.getenv(key)

# Validate secrets
required_keys = [
    "AZURE_OPENAI_ENDPOINT",
    "AZURE_OPENAI_API_KEY",
    "AZURE_OPENAI_DEPLOYMENT"
]
for key in required_keys:
    value = get_secret(key)
    if not value:
        st.error(f"❌ Missing {key} in secrets")
        st.stop()
    # Debug: Show first/last 4 chars only (for security)
    if key == "AZURE_OPENAI_API_KEY":
        masked = f"{value[:4]}...{value[-4:]}" if len(value) > 8 else "***"
        st.sidebar.info(f"🔑 API Key loaded: {masked}")
    elif key == "AZURE_OPENAI_ENDPOINT":
        st.sidebar.info(f"🌐 Endpoint: {value}")
    elif key == "AZURE_OPENAI_DEPLOYMENT":
        st.sidebar.info(f"🚀 Deployment: {value}")

# Initialize Azure OpenAI client
try:
    # Try with the API version from your working local setup
    api_version = get_secret("AZURE_OPENAI_API_VERSION") or "2024-02-01"
    
    client = AzureOpenAI(
        azure_endpoint=get_secret("AZURE_OPENAI_ENDPOINT"),
        api_key=get_secret("AZURE_OPENAI_API_KEY"),
        api_version=api_version
    )
    DEPLOYMENT = get_secret("AZURE_OPENAI_DEPLOYMENT")
    if not DEPLOYMENT:
        st.error("❌ Missing deployment name in secrets")
        st.stop()
except Exception as e:
    st.error(f"❌ Error initializing Azure client: {str(e)}")
    st.stop()

# Constants
RTI_REFERENCE = "TRAOI/R/E/25/00652"
RTI_COMPLAINT_DATE = "12/05/2025"
RTI_APPLICATION_DATE = "09.09.2025"

# ---------- NEW: robust URL extraction ----------
URL_REGEX = re.compile(
    r"""(?ix)
    \b
    (https?://[^\s<>\)\]\}]+)   # capture typical URL up to whitespace or closing punctuation
    """
)

def extract_first_url(text: str) -> str:
    """
    Return the first valid http/https URL in the given text.
    Accepts '•', commas, pipes, newlines, or spaces as separators.
    Filters out 'blob:' and trivial invalid placeholders.
    """
    if not text or str(text).strip() in {"—", "nan", "", "1", "None"}:
        return ""

    s = str(text).strip()

    # Quickly discard blob: type or non-http(s)
    if s.startswith("blob:"):
        return ""

    # Try regex find
    match = URL_REGEX.search(s)
    if not match:
        return ""

    url = match.group(1)

    # Trim trailing punctuation that sometimes sticks to URLs
    url = url.rstrip(".,);]}>\"'")

    # Basic validation: must start with http or https
    if not (url.startswith("http://") or url.startswith("https://")):
        return ""

    return url


def load_knowledge_base():
    """Load and preprocess the knowledge base from Excel file with precise column handling"""
    try:
        # Load both sheets from the Excel file
        sheet1 = pd.read_excel("RTI-Query and Response.xlsx", sheet_name="Sheet1")
        sheet2 = pd.read_excel("RTI-Query and Response.xlsx", sheet_name="Sheet2")

        # Combine both sheets
        combined_df = pd.concat([sheet1, sheet2], ignore_index=True)

        # Clean and standardize column names with precise handling
        clean_columns = []
        for col in combined_df.columns:
            # Handle special characters and whitespace precisely
            clean_col = str(col).strip().replace('\xa0', ' ').replace('`', "'").replace('\u2019', "'")
            clean_columns.append(clean_col)
        combined_df.columns = clean_columns

        # Define the exact column names based on your knowledge base
        theme_col = "Specific type of query (RTI theme)"
        response_col = "TRAI's response (as per letter)"
        url_col = "URLs mentioned in the reply"

        # Verify these columns exist in the dataframe
        missing_columns = []
        for col in [theme_col, response_col, url_col]:
            if col not in combined_df.columns:
                # Try to find a similar column name
                similar_cols = [c for c in combined_df.columns if col.lower() in c.lower()]
                if similar_cols:
                    st.warning(f"Column '{col}' not found exactly. Using '{similar_cols[0]}' instead.")
                    if col == theme_col:
                        theme_col = similar_cols[0]
                    elif col == response_col:
                        response_col = similar_cols[0]
                    elif col == url_col:
                        url_col = similar_cols[0]
                else:
                    missing_columns.append(col)
        if missing_columns:
            st.error(f"❌ Could not find required columns: {', '.join(missing_columns)}")
            st.error(f"Available columns: {', '.join(combined_df.columns)}")
            st.stop()

        # Clean the data precisely
        combined_df = combined_df[~combined_df[url_col].astype(str).str.contains("^1$", case=False, na=False)]
        combined_df = combined_df[~combined_df[url_col].astype(str).str.startswith('blob:')]
        combined_df = combined_df.drop_duplicates(subset=[theme_col], keep='first')
        combined_df = combined_df.dropna(subset=[theme_col, response_col])
        combined_df = combined_df.reset_index(drop=True)

        # Store column mappings
        st.session_state.column_mapping = {
            "theme": theme_col,
            "response": response_col,
            "url": url_col
        }
        return combined_df
    except Exception as e:
        st.error(f"❌ Error loading knowledge base: {str(e)}")
        st.stop()


def split_into_questions(query):
    """Use LLM to intelligently split user query into individual questions"""
    # Clean the query to prevent JSON parsing issues
    cleaned_query = query.replace('"', '\\"').replace('\n', ' ').strip()
    
    prompt = f"""You are an expert assistant for analyzing RTI queries. Your task is to identify and separate individual questions from a user's query.

USER QUERY:
{cleaned_query}

INSTRUCTIONS:
1. Identify all distinct questions in the query
2. Each question should be a complete, standalone question
3. Preserve the original wording as much as possible
4. If the query contains multiple questions (numbered or unnumbered), separate them
5. Return the questions as a JSON array of strings
6. If there's only one question, return an array with one element
7. Remove any numbering from the questions themselves

EXAMPLE INPUT:
1. What is QoS performance? 2. How to file complaints?

EXAMPLE OUTPUT:
["What is QoS performance?", "How to file complaints?"]

Return ONLY the JSON array, nothing else. Do not include any explanatory text."""
    
    try:
        # Call the LLM
        response = client.chat.completions.create(
            model=DEPLOYMENT,
            messages=[
                {"role": "system", "content": "You are a precise question extraction assistant. Return only valid JSON arrays with proper escaping."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=1000
        )
        result = response.choices[0].message.content.strip()
        
        # Clean up the response - remove markdown code blocks if present
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0].strip()
        elif "```" in result:
            result = result.split("```")[1].split("```")[0].strip()
        
        # Remove any leading "json" text
        if result.lower().startswith("json"):
            result = result[4:].strip()
        
        # Try to parse the JSON
        questions = json.loads(result)
        
        # Validate it's a list
        if not isinstance(questions, list):
            st.warning("⚠️ LLM did not return a list. Processing as single query.")
            return [query]
        
        # Validate all elements are strings
        questions = [str(q).strip() for q in questions if q]
        
        if not questions:
            return [query]
            
        return questions
        
    except json.JSONDecodeError as e:
        st.warning(f"⚠️ Could not parse LLM response as JSON. Processing as single query.")
        return [query]
    except Exception as e:
        st.warning(f"⚠️ Could not split questions automatically: {str(e)}. Processing as single query.")
        st.error(f"Full error details: {repr(e)}")  # Show full error for debugging
        return [query]


def find_best_match_with_llm(query, knowledge_base):
    """Use LLM to find the best matching RTI theme for the user's query"""
    # Get column names
    theme_col = st.session_state.column_mapping["theme"]
    response_col = st.session_state.column_mapping["response"]

    # Create a concise knowledge base summary for the LLM
    kb_summary = ""
    themes_list = []
    for idx, row in knowledge_base.iterrows():
        theme = row[theme_col].strip() if isinstance(row[theme_col], str) else str(row[theme_col])
        response = row[response_col].strip() if isinstance(row[response_col], str) else str(row[response_col])
        kb_summary += f"ID: {idx}\nTheme: {theme}\nResponse: {response}\n\n"
        themes_list.append(theme)

    # Create prompt for the LLM
    prompt = f"""You are an expert assistant for TRAI (Telecom Regulatory Authority of India) RTI responses.
Your task is to find the most relevant RTI theme from the knowledge base that matches the user's query.
USER QUERY: "{query}"
KNOWLEDGE BASE (list of RTI themes with IDs):
{kb_summary}
INSTRUCTIONS:
1. Carefully analyze the user query and match it to the most relevant RTI theme from the knowledge base
2. Return ONLY the ID number of the best matching theme (e.g., "3" or "15")
3. If no good match is found, return "-1"
4. Do not explain your reasoning or provide any other text"""
    try:
        # Call the LLM
        response = client.chat.completions.create(
            model=DEPLOYMENT,
            messages=[
                {"role": "system", "content": "You are a precise RTI matching assistant for TRAI."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=10
        )

        # Extract the ID from the response
        match_id = response.choices[0].message.content.strip()

        # Try to parse the ID
        try:
            match_id = int(match_id)
            if 0 <= match_id < len(knowledge_base):
                return knowledge_base.iloc[match_id]
            else:
                return None
        except ValueError:
            # Check if the response contains a theme that matches one in our list
            response_text = response.choices[0].message.content.strip().lower()
            for idx, theme in enumerate(themes_list):
                if theme.lower() in response_text or response_text in theme.lower():
                    return knowledge_base.iloc[idx]
            return None
    except Exception as e:
        st.error(f"❌ Error using LLM for matching: {str(e)}")
        return None


def summarize_response_with_llm(question, response_text, url):
    """Use LLM to convert knowledge base response into natural, professional language"""
    url_instruction = ""
    if url:
        # Encourage natural inclusion of URL
        url_instruction = (
            f"\n\nIMPORTANT: The URL is: {url}\n"
            f"You MUST include this URL in your response naturally at the end. "
            f"Use phrases like 'Details are available at {url}' or "
            f"'Please refer to {url}' or 'The information can be accessed at {url}'"
        )
    else:
        url_instruction = "\n\nNOTE: No URL is available. Do NOT mention anything about URLs or links in your response."

    prompt = f"""You are a professional RTI response writer for TRAI. Your task is to convert abbreviated/technical knowledge base entries into clear, professional, natural-sounding responses.
QUESTION: {question}
KNOWLEDGE BASE ENTRY: {response_text}
{url_instruction}
INSTRUCTIONS:
1. Convert the abbreviated entry into 2-4 complete, professional sentences
2. Write in formal government communication style (similar to official letters)
3. Be specific and factual - don't add information not in the knowledge base entry
4. Make it sound like a human government official wrote it, not a database entry
5. Do NOT start with "The information..." or "TRAI informs that..." - just provide the direct answer
6. Keep it concise but complete
7. Do NOT mention "No URL is available" or similar phrases - just omit URL mention if none exists
Return ONLY the natural response, nothing else."""
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT,
            messages=[
                {"role": "system", "content": "You are a professional RTI response writer. Write clear, natural, formal responses."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=300
        )
        llm_response = response.choices[0].message.content.strip()

        # Double-check: If URL exists but LLM didn't include it, append it manually (markdown link for Streamlit)
        if url and (url not in llm_response):
            llm_response += f" Relevant details are available at: {url}"

        return llm_response
    except Exception:
        # Fallback to original if LLM fails
        if url:
            return f"{response_text} Relevant details are available at: {url}"
        return response_text


def generate_multi_response(questions_and_matches):
    """Generate the formatted response for multiple questions - ONLY from knowledge base"""
    # Get column names
    response_col = st.session_state.column_mapping["response"]
    url_col = st.session_state.column_mapping["url"]

    # Build the header
    response = f"""Sir,

Please refer to your RTI application dated {RTI_APPLICATION_DATE} filed online on the RTI portal vide Reg. No. {RTI_REFERENCE}, for providing information under the provisions of the RTI Act 2005. In this context, the following is furnished: -

---

**Information Sought:**

"""
    
    # Add the questions with proper line breaks
    for idx, (question, _) in enumerate(questions_and_matches, 1):
        response += f"{idx}. {question}\n\n"
    
    response += "---\n\n"

    # Start the table with proper formatting
    response += "| S.No. | Reply |\n"
    response += "|:------|:------|\n"

    # Add answers ONLY from knowledge base - no static responses
    for idx, (question, match) in enumerate(questions_and_matches, 1):
        if match is not None:
            response_text = match[response_col]
            raw_urls = str(match[url_col]).strip()

            # Use regex-based extraction
            url = extract_first_url(raw_urls)

            # Use LLM to make response natural and professional
            natural_response = summarize_response_with_llm(question, response_text, url)

            # Replace newlines in the response with <br> for proper table formatting
            response_text_final = natural_response.replace('\n', '<br>')
        else:
            response_text_final = (
                "The information sought could not be precisely matched in TRAI's knowledge base. "
                "Please consider rephrasing your query or contacting the relevant department directly."
            )

        response += f"| {idx}. | {response_text_final} |\n"

    return response


def create_response_excel(response_text, questions_and_matches):
    """Create an Excel file with the formatted response"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font

    # Get column names
    response_col = st.session_state.column_mapping["response"]
    url_col = st.session_state.column_mapping["url"]

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "RTI Response"

    # Add header lines
    ws.append(["Sir,"])
    ws.append([f"Please refer to your RTI application dated {RTI_APPLICATION_DATE} filed online on the RTI portal vide Reg. No. {RTI_REFERENCE}, for providing information under the provisions of the RTI Act 2005. In this context, the following is furnished: -"])
    ws.append([""])

    # Add "Information Sought" section
    ws.append(["Information Sought:"])
    ws.append([""])
    for idx, (question, _) in enumerate(questions_and_matches, 1):
        ws.append([f"{idx}. {question}"])
        ws.append([""])

    # Add table headers
    ws.append(["S.No.", "Reply"])
    header_row = ws.max_row

    # Add answers ONLY from knowledge base
    for idx, (question, match) in enumerate(questions_and_matches, 1):
        if match is not None:
            response_text = match[response_col]
            raw_urls = str(match[url_col]).strip()

            # Use regex-based extraction
            url = extract_first_url(raw_urls)

            # Use LLM to make response natural and professional
            natural_response = summarize_response_with_llm(question, response_text, url)

            # For Excel, append URL explicitly at the end so it's visible/clickable
            if url and (url not in natural_response):
                answer = f"{natural_response} Relevant details are available at: {url}"
            else:
                answer = natural_response
        else:
            answer = (
                "The information sought could not be precisely matched in TRAI's knowledge base. "
                "Please consider rephrasing your query or contacting the relevant department directly."
            )

        ws.append([f"{idx}.", answer])

    # Apply borders to the table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply formatting to all cells from header row onwards
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Make header row bold
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 100

    # Save to stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def main():
    st.set_page_config(
        page_title="TRAI RTI Response Generator",
        page_icon="📄",
        layout="wide"
    )
    
    st.title("📄 TRAI RTI Response Generator")
    st.markdown("### Knowledge Base Powered Response System")
    st.markdown("---")

    # Load knowledge base
    knowledge_base = load_knowledge_base()

    # Sidebar with improved styling
    with st.sidebar:
        st.success(f"✅ **Knowledge Base Loaded**")
        st.metric(label="RTI Themes Available", value=len(knowledge_base))
        
        st.markdown("---")
        st.info("💡 **Multi-Question Support**\n\nAsk multiple questions in a single submission. Each answer comes directly from the knowledge base.")
        
        st.markdown("---")
        st.warning("⚠️ **Knowledge Base Only**\n\nAll answers are matched from your Excel knowledge base only.")

    # User input with improved styling
    st.subheader("📝 Submit Your RTI Query")
    
    user_query = st.text_area(
        "Enter your RTI query (Information sought):",
        height=200,
        placeholder="""You can ask single or multiple questions. Examples:

Single question:
Copies of QoS compliance reports submitted by Airtel for the last three years

Multiple questions (numbered):
1. Copies of QoS compliance reports submitted by Airtel for the last three years
2. Number of consumer complaints received regarding poor internet quality
3. Tariff plan filings with TRAI for prepaid mobile services

Multiple questions (unnumbered):
What are the QoS parameters? How do I file a complaint? Where can I find tariff information?"""
    )

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        generate_button = st.button("🚀 Generate Response", type="primary", disabled=not user_query, use_container_width=True)

    if generate_button:
        with st.spinner("🔄 Processing your query with Azure OpenAI LLM..."):
            # Step 1: Split the query into individual questions
            questions = split_into_questions(user_query)
            st.info(f"🔍 Detected **{len(questions)}** question(s) in your query")

            # Step 2: Find matches for each question
            questions_and_matches = []
            
            # Create a progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for idx, question in enumerate(questions, 1):
                status_text.text(f"Processing question {idx}/{len(questions)}...")
                progress_bar.progress(idx / len(questions))
                
                match = find_best_match_with_llm(question, knowledge_base)
                questions_and_matches.append((question, match))
                
                if match is not None:
                    theme_col = st.session_state.column_mapping["theme"]
                    st.success(f"✅ **Question {idx}** matched to: *{match[theme_col]}*")
                else:
                    st.warning(f"⚠️ **Question {idx}**: No match found in knowledge base")
            
            progress_bar.empty()
            status_text.empty()

            # Step 3: Generate the complete response
            response_text = generate_multi_response(questions_and_matches)

            # Display the response with improved styling
            st.markdown("---")
            st.subheader("📋 Generated Response")
            
            # Use a container with custom styling
            with st.container():
                st.markdown(response_text, unsafe_allow_html=True)

            st.markdown("---")

            # Create Excel file
            excel_file = create_response_excel(response_text, questions_and_matches)

            # Provide download button with better styling
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 Download Response as Excel",
                    data=excel_file,
                    file_name="RTI_Response.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

if __name__ == "__main__":
    main()
